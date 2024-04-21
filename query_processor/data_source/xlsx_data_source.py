from typing import Callable

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from query_processor.data_source.source import DataSource


class XlsxDataSource(DataSource):
    _fields_to_xlsx_column = {}

    def __init__(self,
                 xlsx_file_path,
                 worksheet_name: str | None = None,
                 min_row: int | None = 1,
                 max_row: int | None = None,
                 left_col: int | None = None,
                 right_row: int | None = None,
                 skip_head: bool = True):
        self.skip_head = skip_head
        self.right_row = right_row
        self.left_col = left_col
        self.max_row = max_row
        self.min_row = min_row
        self.worksheet_name = worksheet_name
        self.xlsx_file_path = xlsx_file_path

    def parse_row(self, row: tuple[Cell, ...]) -> dict:
        res = {}
        for field, column in self._fields_to_xlsx_column.items():
            res[field] = row[column - 1].value
        return res

    def filter_row(self, row: dict) -> bool:
        pass

    def get(self, filter_dict: dict | None = None,
            filter_predicate: Callable[[dict], bool] | None = None,
            curr_min_row: int | None = None,
            curr_max_row: int | None = None,
            all_fields: bool = False) -> list[dict]:
        """
        :param filter_dict: dict Словарь, где ключ - название столбца,
        а значение - значение столбца или список значений.

        Фильтры применяются через ИЛИ !!!
        То есть если строка подойдёт хотя бы под один фильтр, то эта строка будет выбрана.

        :param filter_predicate: функция, проверяющая строку для фильтрации

        :param curr_min_row: текущий минимальный индекс строки, заменят собой self.min_row

        :param curr_max_row: текущий максимальный индекс строки, заменят собой self.max_row

        :param all_fields: флаг, показывающий, что нужно вернуть все строки, без фильтрации

        :return: Список строк прошедших фильтрацию
        """
        workbook = openpyxl.load_workbook(self.xlsx_file_path)
        worksheet: Worksheet = workbook[self.worksheet_name] if self.worksheet_name is not None else workbook.active
        min_row = self.min_row
        if self.skip_head and self.min_row is not None:
            min_row += 1
        res: list[dict] = []
        for row in worksheet.iter_rows(min_row=min_row, max_row=self.max_row, min_col=self.left_col,
                                       max_col=self.right_row):
            parsed_row = self.parse_row(row)
            if all_fields:
                res.append(parsed_row)
                continue
            if filter_dict is not None:
                for field, val in filter_dict.items():
                    if field in parsed_row:
                        if isinstance(filter_dict[field], list):
                            if isinstance(filter_dict[field][0], str):
                                if str(parsed_row[field]) in filter_dict[field]:
                                    res.append(parsed_row)
                            elif isinstance(filter_dict[field][0], int):
                                if parsed_row[field] in filter_dict[field]:
                                    res.append(parsed_row)
                            continue
                        elif str(parsed_row[field]) == str(val):
                            res.append(parsed_row)
            if self.filter_row(parsed_row):
                res.append(parsed_row)
            if filter_predicate is not None:
                if filter_predicate(parsed_row):
                    res.append(parsed_row)
        return res
